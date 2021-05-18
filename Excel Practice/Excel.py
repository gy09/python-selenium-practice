from openpyxl import load_workbook, cell
import openpyxl


class Excel_Test:
    def test_excel(self):
        # creating a new excel file
        wb = openpyxl.Workbook()
        # creating sheet in the excel
        sheet = wb.create_sheet("Dummy Sheet")
        # entering the value in the excel
        sheet["A1"].value = "hello excel"

        sheet["B2"].value = "hell yeah"

        wb.save("Test_excel.xlxs")

    def retrieve_excel(self):
        filepath = "Test_excel.xlxs"
        wb = load_workbook(filepath)

        sheet = wb.active
        a1 = sheet["A1"]
        print(a1.value)
        b2 = sheet.cell(row=2, column='B')
        print(b2.value)


obj = Excel_Test()
obj.test_excel()
obj.retrieve_excel()
